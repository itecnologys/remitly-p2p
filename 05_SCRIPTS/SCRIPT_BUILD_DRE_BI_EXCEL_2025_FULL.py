"""
DRE 100 LPs — **Power BI / Excel (ano-calendário 2025)**.

- **Uma linha por rematch** (Dock reload R$ 2,00): total **84.107** linhas (= R$ 168.214,40 / 2).
- **GTV** por linha soma **exactamente** R$ 1.409.813.152,46 (DRE `INVESTOR_DRE_CASE_STUDY_100.md`).
- Cada LP-dia **activo**: **3–6 ciclos**; rematches desse dia repartem-se por ciclos (composição aleatória ≥1 por ciclo).
- **Macro hash:** uma por **data** com transacção (todas as linhas do dia partilham `macro_hash`).
- **Minihash:** uma por rematch; coluna `vinculo_minihash_macrohash`.
- **PTAX USD/BRL:** BCB Olinda `CotacaoDolarPeriodo` (01/01–31/12/2025); *carry-forward* para dias sem cotação.

Saídas (`06_DATA/`):
  DATA_DRE_BI_PTAX_2025.json
  DATA_DRE_BI_LP_MASTER_2025.json
  DATA_DRE_BI_REMATCHES_2025.csv
  DATA_DRE_BI_WORKBOOK_2025.xlsx
"""
from __future__ import annotations

import csv
import hashlib
import json
import random
import urllib.error
import urllib.request
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]

# --- Totais DRE (Sincronização CET 2,47% / Spread LP 1,08%) ---
GTV_TOTAL_BRL = 1409813152.46
DOCK_RELOAD_TOTAL = 168214.40
DOCK_PER_REMATCH = 2.0
N_REMATCHES = int(round(DOCK_RELOAD_TOTAL / DOCK_PER_REMATCH)) 

# Regras de Tiers (CET e Spread) para o DRE 2025
TIERS = [
    {"limit_usd": 300, "cet": 0.0247, "spread_lp": 0.0108},
    {"limit_usd": 500, "cet": 0.0215, "spread_lp": 0.0095},
    {"limit_usd": 1000, "cet": 0.0185, "spread_lp": 0.0080},
    {"limit_usd": float('inf'), "cet": 0.0150, "spread_lp": 0.0060}
]

IOF_RATE = 0.0038

OUT_PTAX = REPO / "06_DATA" / "DATA_DRE_BI_PTAX_2025.json"
OUT_LP = REPO / "06_DATA" / "DATA_DRE_BI_LP_MASTER_2025.json"
OUT_CSV = REPO / "06_DATA" / "DATA_DRE_BI_REMATCHES_2025.csv"
OUT_XLSX = REPO / "06_DATA" / "DATA_DRE_BI_WORKBOOK_2025.xlsx"

PRIMEIROS_M = (
    "Carlos João Lucas Rafael Bruno Felipe Gustavo André Rodrigo Thiago Daniel Eduardo "
    "Gabriel Henrique Igor José Leonardo Marcelo Nelson Otávio Paulo Pedro Samuel "
    "Vinícius Wagner Xavier"
).split()
PRIMEIROS_F = (
    "Ana Maria Juliana Fernanda Patrícia Camila Beatriz Larissa Amanda Renata "
    "Adriana Bianca Carla Daniela Elisa Fabiana Gabriela Helena Isabela Júlia "
    "Karina Letícia Mariana Natália Paula Roberta Sandra Tatiana Vanessa"
).split()
SOBRENOMES = (
    "Silva Santos Oliveira Souza Lima Rodrigues Ferreira Alves Pereira Costa "
    "Martins Barbosa Rocha Dias Gomes Carvalho Araújo Melo Cardoso Teixeira "
    "Correia Monteiro Cavalcanti Nascimento Freitas Moreira Duarte Pinto Ramos"
).split()
CIDADES = [
    ("São Paulo", "SP"),
    ("Rio de Janeiro", "RJ"),
    ("Belo Horizonte", "MG"),
    ("Curitiba", "PR"),
    ("Porto Alegre", "RS"),
    ("Salvador", "BA"),
    ("Fortaleza", "CE"),
    ("Recife", "PE"),
    ("Brasília", "DF"),
    ("Manaus", "AM"),
    ("Goiânia", "GO"),
    ("Campinas", "SP"),
    ("Florianópolis", "SC"),
    ("Belém", "PA"),
    ("Natal", "RN"),
    ("Vitória", "ES"),
    ("João Pessoa", "PB"),
    ("Maceió", "AL"),
    ("Aracaju", "SE"),
    ("Cuiabá", "MT"),
]


@dataclass
class LPRow:
    lp_id: str
    nome_completo: str
    genero: str
    cidade: str
    estado: str


@dataclass
class PTAXRow:
    data: str
    timestamp_ptax_oficial: str
    cotacao_compra: float
    cotacao_venda: float
    usd_brl_mid: float


def fetch_ptax_2025() -> tuple[list[PTAXRow], str]:
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
        "CotacaoDolarPeriodo(dataInicial=@di,dataFinalCotacao=@df)"
        "?@di='01-01-2025'&@df='12-31-2025'&$format=json&$orderby=dataHoraCotacao%20asc"
    )
    with urllib.request.urlopen(url, timeout=90) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    rows = data.get("value", [])
    out: list[PTAXRow] = []
    for row in rows:
        dh = str(row.get("dataHoraCotacao", ""))
        day = dh[:10]
        c, v = float(row["cotacaoCompra"]), float(row["cotacaoVenda"])
        out.append(
            PTAXRow(
                data=day,
                timestamp_ptax_oficial=dh,
                cotacao_compra=c,
                cotacao_venda=v,
                usd_brl_mid=round((c + v) / 2.0, 4),
            )
        )
    return out, "BCB Olinda — PTAX CotacaoDolarPeriodo (2025)"


def carry_forward_ptax(rows: list[PTAXRow]) -> dict[str, PTAXRow]:
    """Todos os dias 2025-01-01..2025-12-31 → última PTAX publicada até esse dia (incl. 1/1 com 1ª cotação)."""
    by_pub: dict[str, PTAXRow] = {}
    for r in rows:
        by_pub[r.data] = r
    first = min(by_pub.keys())
    cur: PTAXRow = by_pub[first]
    out: dict[str, PTAXRow] = {}
    d0 = date(2025, 1, 1)
    for i in range(365):
        d = d0 + timedelta(days=i)
        key = d.isoformat()
        if key in by_pub:
            cur = by_pub[key]
        out[key] = cur
    return out


def build_lps(rng: random.Random) -> list[LPRow]:
    rows: list[LPRow] = []
    for i in range(1, 101):
        gen = "F" if rng.random() < 0.52 else "M"
        p1 = rng.choice(PRIMEIROS_F if gen == "F" else PRIMEIROS_M)
        s1, s2 = rng.choice(SOBRENOMES), rng.choice(SOBRENOMES)
        if s2 == s1:
            s2 = SOBRENOMES[(SOBRENOMES.index(s1) + 3) % len(SOBRENOMES)]
        nome = f"{p1} {s1} {s2}"
        city, uf = rng.choice(CIDADES)
        rows.append(LPRow(lp_id=f"LP_{i:03d}", nome_completo=nome, genero=gen, cidade=city, estado=uf))
    return rows


def minihash(seed: str) -> str:
    return "mh0x" + hashlib.sha256(seed.encode()).hexdigest()[:56]


def macro_hash(day_key: str, ptax_ts: str) -> str:
    return "MH0x" + hashlib.sha256(f"MACRO|{day_key}|{ptax_ts}".encode()).hexdigest()[:56]


def random_composition(total: int, parts: int, rng: random.Random) -> list[int]:
    """Parte `total` em `parts` inteiros ≥1."""
    if parts == 1:
        return [total]
    cuts = sorted(rng.sample(range(1, total), parts - 1))
    out = []
    prev = 0
    for c in cuts + [total]:
        out.append(c - prev)
        prev = c
    return out


def allocate_slots(R: int, rng: random.Random) -> tuple[list[dict], list[tuple[int, int]]]:
    """
    Retorna lista de slots activos: {lp_index 0..99, day_index 0..364, cycles, rematches, comp (lista por ciclo)}.
    """
    Kmin = (R + 5) // 6
    Kmax = R // 3
    pairs = [(lp, d) for lp in range(100) for d in range(365)]
    rng.shuffle(pairs)

    for attempt in range(5000):
        K = rng.randint(Kmin, Kmax)
        cycles = [rng.randint(3, 6) for _ in range(K)]
        S = sum(cycles)
        while S > R:
            j = max(range(K), key=lambda x: cycles[x])
            if cycles[j] > 3:
                cycles[j] -= 1
                S -= 1
            else:
                break
        if S > R:
            continue
        Q = R - S
        extras = [0] * K
        for _ in range(Q):
            extras[rng.randrange(K)] += 1
        rem = [cycles[i] + extras[i] for i in range(K)]
        if sum(rem) != R:
            continue
        slots = []
        used_pairs = pairs[:K]
        for idx in range(K):
            lp_i, d_i = used_pairs[idx]
            M = rem[idx]
            # 3–6 ciclos por LP-dia activo, nunca superiores ao número de rematches desse dia
            cyc = rng.randint(3, min(6, M))
            comp = random_composition(M, cyc, rng)
            slots.append(
                {
                    "lp_index": lp_i,
                    "day_index": d_i,
                    "cycles": cyc,
                    "rematches": M,
                    "composition": comp,
                }
            )
        return slots, used_pairs

    raise RuntimeError("Falha ao alocar rematches; ajustar RNG ou limites.")


def distribute_gtv_exact(n: int, total: float, rng: random.Random) -> list[float]:
    """Distribui `total` em `n` partes positivas com soma exacta (2 casas)."""
    cents = int(round(total * 100))
    w = [rng.random() for _ in range(n)]
    s = sum(w)
    parts = [int(cents * x / s) for x in w]
    diff = cents - sum(parts)
    order = sorted(range(n), key=lambda i: (w[i] * cents / s) - parts[i], reverse=True)
    for k in range(diff):
        parts[order[k % n]] += 1
    return [p / 100.0 for p in parts]


def distribute_exact(total: float, weights: list[float]) -> list[float]:
    cents = int(round(total * 100))
    s = sum(weights)
    parts = [int(cents * w / s) for w in weights]
    diff = cents - sum(parts)
    order = sorted(range(len(weights)), key=lambda i: (weights[i] * cents / s) - parts[i], reverse=True)
    for k in range(diff):
        parts[order[k % len(weights)]] += 1
    return [p / 100.0 for p in parts]


def main() -> None:
    rng = random.Random(20250218)
    ptax_rows, src = fetch_ptax_2025()
    ptax_day = carry_forward_ptax(ptax_rows)
    OUT_PTAX.parent.mkdir(parents=True, exist_ok=True)
    OUT_PTAX.write_text(
        json.dumps(
            {
                "fonte": src,
                "registos_publicacao": [asdict(x) for x in ptax_rows],
                "mapa_dia_util_ptax_carryforward": {k: asdict(v) for k, v in sorted(ptax_day.items())},
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    lps = build_lps(rng)
    OUT_LP.write_text(json.dumps([asdict(x) for x in lps], indent=2, ensure_ascii=False), encoding="utf-8")

    slots, _ = allocate_slots(N_REMATCHES, rng)

    # achatar em lista de (slot, cycle_id, rem_in_cycle) com contagem de rematches por ciclo
    flat_slots: list[tuple[dict, int, int]] = []
    for sl in slots:
        cid = 0
        for c_idx, k_rem in enumerate(sl["composition"], start=1):
            for _ in range(k_rem):
                flat_slots.append((sl, c_idx, cid))
                cid += 1

    assert len(flat_slots) == N_REMATCHES

    gtv_parts = distribute_gtv_exact(N_REMATCHES, GTV_TOTAL_BRL, rng)
    assert abs(sum(gtv_parts) - GTV_TOTAL_BRL) < 0.01

    # Cálculo dinâmico linha a linha com base em Tiers de CET
    start = date(2025, 1, 1)
    rows_out: list[dict] = []
    
    # Pesos uniformes para o Dock (R$ 2,00 por linha)
    dock_line = distribute_exact(DOCK_RELOAD_TOTAL, [1.0] * N_REMATCHES)

    for idx, (sl, cycle_id, _rid) in enumerate(flat_slots):
        lp_i = sl["lp_index"]
        di = sl["day_index"]
        d = start + timedelta(days=di)
        d_key = d.isoformat()
        pr = ptax_day.get(d_key)
        if pr is None:
            raise KeyError(f"Sem PTAX carry-forward para {d_key}")
        mh = macro_hash(d_key, pr.timestamp_ptax_oficial)
        gtv = round(gtv_parts[idx], 2)
        
        # Lógica de Tiers Dinâmicos baseada em GTV USD
        gtv_usd = gtv / pr.usd_brl_mid
        tier = next(t for t in TIERS if gtv_usd <= t["limit_usd"])
        
        cet_rate = tier["cet"]
        lp_rate = tier["spread_lp"]
        
        # Decomposição CET: Spread LP + Orchestration + IOF
        fee_orch_rate = cet_rate - lp_rate - IOF_RATE
        
        spread_lp_brl = round(gtv * lp_rate, 2)
        iof_brl = round(gtv * IOF_RATE, 2)
        fluxus_orch_brl = round(gtv * fee_orch_rate, 2)
        cet_total_brl = round(gtv * cet_rate, 2)

        seed = f"{d_key}|{lps[lp_i].lp_id}|{idx}|C{cycle_id}|REM"
        mi = minihash(seed)
        ts = datetime.combine(d, datetime.min.time()).replace(hour=8 + (idx % 14), minute=(idx * 17) % 60)
        ts_iso = ts.strftime("%Y-%m-%dT%H:%M:%S-03:00")
        rows_out.append(
            {
                "rematch_id": f"RM-{idx+1:06d}",
                "data_operacao": d_key,
                "timestamp_remessa_iso": ts_iso,
                "timestamp_ptax_bcb_iso": pr.timestamp_ptax_oficial,
                "usd_brl_ptax_compra": pr.cotacao_compra,
                "usd_brl_ptax_venda": pr.cotacao_venda,
                "usd_brl_ptax_mid": pr.usd_brl_mid,
                "macro_hash": mh,
                "minihash": mi,
                "vinculo_minihash_macrohash": f"{mi}→{mh}",
                "lp_id": lps[lp_i].lp_id,
                "lp_nome": lps[lp_i].nome_completo,
                "lp_genero": lps[lp_i].genero,
                "lp_cidade": lps[lp_i].cidade,
                "lp_estado": lps[lp_i].estado,
                "ciclos_no_dia_lp": sl["cycles"],
                "ciclo_id": cycle_id,
                "ciclos_no_dia_lp": sl["cycles"],
                "ciclo_id": cycle_id,
                "dock_reload_brl": round(dock_line[idx], 2),
                "gtv_brl": gtv,
                "gtv_usd_equiv_ptax_mid": round(gtv_usd, 2),
                "cet_total_brl": cet_total_brl,
                "lp_spread_brl": spread_lp_brl,
                "iof_recolhido_brl": iof_brl,
                "fluxus_orchestration_fee_brl": fluxus_orch_brl,
                "stability_pool_alloc_2pct_linha": round(gtv * 0.02, 2),
            }
        )

    # verificações
    s_gtv = sum(r["gtv_brl"] for r in rows_out)
    s_dock = sum(r["dock_reload_brl"] for r in rows_out)
    s_orch = sum(r["fluxus_orchestration_fee_brl"] for r in rows_out)
    s_stab = sum(r["stability_pool_alloc_2pct_linha"] for r in rows_out)
    s_lp = sum(r["lp_spread_brl"] for r in rows_out)
    
    print(f"--- RESULTADO AUDITORIA EXCEL BI ---")
    print(f"GTV Total: R$ {s_gtv:,.2f}")
    print(f"Receita Bruta Fluxus (Orch): R$ {s_orch:,.2f} ({ (s_orch/s_gtv)*100 :.2f}%)")
    print(f"Lucro Bruto LPs: R$ {s_lp:,.2f}")
    print(f"Stability Pool: R$ {s_stab:,.2f}")

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader()
        w.writerows(rows_out)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fill_h = PatternFill("solid", fgColor="1B365D")
    font_h = Font(color="FFFFFF", bold=True)

    ws0 = wb.active
    ws0.title = "REFERENCIA"
    meta = [
        ("GTV total (soma linhas)", f"{s_gtv:.2f}"),
        ("Receita Fluxus (Orc)", f"{s_orch:.2f}"),
        ("Lucro LPs total", f"{s_lp:.2f}"),
        ("Stability Pool total", f"{s_stab:.2f}"),
        ("Rematches", str(N_REMATCHES)),
        ("Dock reload total", f"{s_dock:.2f}"),
        ("Ciclos por LP-dia activo", "3 a 6 (aleatório reprodutível)"),
        ("Macro hash", "Por dia com operação (PTAX BCB carry-forward)"),
        ("Fonte PTAX", src),
    ]
    for i, (a, b) in enumerate(meta, 1):
        ws0.cell(i, 1, a).font = Font(bold=True)
        ws0.cell(i, 2, b)

    ws1 = wb.create_sheet("PTAX_DIARIO")
    ws1.append(["data", "timestamp_ptax_oficial", "cotacao_compra", "cotacao_venda", "usd_brl_mid"])
    for c in ws1[1]:
        c.fill = fill_h
        c.font = font_h
    for dk in sorted(ptax_day.keys()):
        r = ptax_day[dk]
        ws1.append([dk, r.timestamp_ptax_oficial, r.cotacao_compra, r.cotacao_venda, r.usd_brl_mid])

    ws2 = wb.create_sheet("LP_MASTER")
    ws2.append(["lp_id", "nome_completo", "genero", "cidade", "estado"])
    for c in ws2[1]:
        c.fill = fill_h
        c.font = font_h
    for lp in lps:
        ws2.append([lp.lp_id, lp.nome_completo, lp.genero, lp.cidade, lp.estado])

    ws3 = wb.create_sheet("REMATCHES")
    keys = list(rows_out[0].keys())
    ws3.append(keys)
    for c in ws3[1]:
        c.fill = fill_h
        c.font = font_h
    money = {
        "gtv_brl",
        "dock_reload_brl",
        "cet_total_brl",
        "lp_spread_brl",
        "iof_recolhido_brl",
        "fluxus_orchestration_fee_brl",
        "stability_pool_alloc_2pct_linha",
        "gtv_usd_equiv_ptax_mid",
        "usd_brl_ptax_mid",
    }
    hashes = {"minihash", "macro_hash", "vinculo_minihash_macrohash"}
    col_idx = {k: i + 1 for i, k in enumerate(keys)}
    mf = PatternFill("solid", fgColor="FFF2CC")
    hf = PatternFill("solid", fgColor="C6E0B4")
    for row in rows_out:
        ws3.append([row[k] for k in keys])
    # destaque só nas primeiras 2000 linhas para não explodir tempo de gravação
    lim = min(ws3.max_row, 2000)
    for name, fl in ((money, mf), (hashes, hf)):
        for k in name:
            if k not in col_idx:
                continue
            L = get_column_letter(col_idx[k])
            for rr in range(2, lim + 1):
                ws3[f"{L}{rr}"].fill = fl

    wb.save(OUT_XLSX)
    print("OK", OUT_XLSX)
    print("GTV check", s_gtv, GTV_TOTAL_BRL)
    print("rematches", len(rows_out))


if __name__ == "__main__":
    main()
