"""
Gera Excel + JSON/CSV para Power BI: DRE 100 LPs — **cenário novembro/2025**.

- **USD/BRL:** API oficial BCB Olinda `CotacaoDolarPeriodo` (PTAX compra/venda → média).
  Fallback: série diária pública (PoundSterlingLive, médias USD-BRL em ### November 2025).
- **Minihash / Macro Hash:** cada ciclo tem `minihash` determinístico; `macro_hash` ancora o
  bucket diário (agregação conforme blueprint: minihashes → macro hash temporal).
- **LPs:** 100 nomes plausíveis BR + género + cidade/UF.

Saídas em `06_DATA/`:
  DATA_DRE_BI_LP_MASTER_NOV2025.json
  DATA_DRE_BI_USD_BRL_PTAX_NOV2025.json
  DATA_DRE_BI_TRANSACTIONS_SAMPLE_NOV2025.csv
  DATA_DRE_BI_WORKBOOK_NOV2025.xlsx
"""
from __future__ import annotations

import csv
import hashlib
import json
import random
import urllib.error
import urllib.request
from dataclasses import dataclass, asdict
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
OUT_JSON_LP = REPO / "06_DATA" / "DATA_DRE_BI_LP_MASTER_NOV2025.json"
OUT_JSON_FX = REPO / "06_DATA" / "DATA_DRE_BI_USD_BRL_PTAX_NOV2025.json"
OUT_CSV_TX = REPO / "06_DATA" / "DATA_DRE_BI_TRANSACTIONS_SAMPLE_NOV2025.csv"
OUT_XLSX = REPO / "06_DATA" / "DATA_DRE_BI_WORKBOOK_HARMONIZED_2026.xlsx"

# Fallback: 20 dias úteis nov/2025 → médias na secção "### November" (PoundSterlingLive USD-BRL 2025).
FALLBACK_DATES = [
    "2025-11-03",
    "2025-11-04",
    "2025-11-05",
    "2025-11-06",
    "2025-11-07",
    "2025-11-10",
    "2025-11-11",
    "2025-11-12",
    "2025-11-13",
    "2025-11-14",
    "2025-11-17",
    "2025-11-18",
    "2025-11-19",
    "2025-11-20",
    "2025-11-21",
    "2025-11-24",
    "2025-11-25",
    "2025-11-26",
    "2025-11-27",
    "2025-11-28",
]
FALLBACK_MID = [
    5.3576,
    5.3357,
    5.3837,
    5.3216,
    5.3526,
    5.3305,
    5.3561,
    5.3864,
    5.3250,
    5.3557,
    5.3844,
    5.3325,
    5.3904,
    5.3299,
    5.3602,
    5.3884,
    5.3823,
    5.4131,
    5.3562,
    5.3846,
]

PRIMEIROS_M = (
    "Carlos João Lucas Rafael Bruno Felipe Gustavo André Rodrigo Thiago Daniel Eduardo "
    "Gabriel Henrique Igor José Leonardo Marcelo Nelson Otávio Paulo Pedro Samuel "
    "Vinícius Wagner Xavier"
).split()
PRIMEIROS_F = (
    "Ana Maria Juliana Fernanda Patrícia Camila Beatriz Larissa Amanda Renata "
    "Adriana Bianca Carla Daniela Elisa Fabiana Gabriela Helena Isabela Júlia "
    "Karina Letícia Mariana Natália Paula Roberta Sandra Tatiana Vanessa"
).split()
SOBRENOMES = (
    "Silva Santos Oliveira Souza Lima Rodrigues Ferreira Alves Pereira Costa "
    "Martins Barbosa Rocha Dias Gomes Carvalho Araújo Melo Cardoso Teixeira "
    "Correia Monteiro Cavalcanti Nascimento Freitas Moreira Duarte Pinto Ramos"
).split()
CIDADES = [
    ("São Paulo", "SP"),
    ("Rio de Janeiro", "RJ"),
    ("Belo Horizonte", "MG"),
    ("Curitiba", "PR"),
    ("Porto Alegre", "RS"),
    ("Salvador", "BA"),
    ("Fortaleza", "CE"),
    ("Recife", "PE"),
    ("Brasília", "DF"),
    ("Manaus", "AM"),
    ("Goiânia", "GO"),
    ("Campinas", "SP"),
    ("Florianópolis", "SC"),
    ("Belém", "PA"),
    ("Natal", "RN"),
    ("Vitória", "ES"),
    ("João Pessoa", "PB"),
    ("Maceió", "AL"),
    ("Aracaju", "SE"),
    ("Cuiabá", "MT"),
]


@dataclass
class LPRow:
    lp_id: str
    nome_completo: str
    genero: str
    cidade: str
    estado: str
    capital_inicial: float
    current_capital: float


@dataclass
class PTAXRow:
    data: str
    timestamp_ptax_oficial: str
    cotacao_compra: float
    cotacao_venda: float
    usd_brl_mid: float


def fetch_bcb_ptax() -> tuple[list[PTAXRow], str] | tuple[None, None]:
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
        "CotacaoDolarPeriodo(dataInicial=@di,dataFinalCotacao=@df)"
        "?@di='11-01-2025'&@df='11-30-2025'&$format=json&$orderby=dataHoraCotacao%20asc"
    )
    try:
        with urllib.request.urlopen(url, timeout=35) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError):
        return None, None
    rows = data.get("value")
    if not rows:
        return None, None
    out: list[PTAXRow] = []
    for row in rows:
        dh = str(row.get("dataHoraCotacao", ""))
        day = dh[:10]
        c, v = float(row["cotacaoCompra"]), float(row["cotacaoVenda"])
        out.append(
            PTAXRow(
                data=day,
                timestamp_ptax_oficial=dh,
                cotacao_compra=c,
                cotacao_venda=v,
                usd_brl_mid=round((c + v) / 2.0, 4),
            )
        )
    return out, "BCB Olinda — PTAX CotacaoDolarPeriodo (nov/2025)"


def get_fluxus_spread_pct(gtv_usd: float) -> float:
    v = float(gtv_usd or 0)
    if v <= 0: return 0.0
    if v <= 100: return 0.0108
    if v <= 1000:
        t = (v - 100) / 900
        return 0.0108 + t * (0.0073 - 0.0108)
    if v <= 5000:
        t = (v - 1000) / 4000
        return 0.0073 + t * (0.0060 - 0.0073)
    if v <= 20000:
        t = (v - 5000) / 15000
        return 0.0060 + t * (0.0055 - 0.0060)
    if v <= 100000:
        t = (v - 20000) / 80000
        return 0.0055 + t * (0.0050 - 0.0055)
    return 0.0050


def fallback_ptax() -> tuple[list[PTAXRow], str]:
    fake_ts = "13:05:00.000"
    rows = []
    for d, mid in zip(FALLBACK_DATES, FALLBACK_MID, strict=True):
        spread = 0.0003
        c = round(mid - spread / 2, 4)
        v = round(mid + spread / 2, 4)
        rows.append(
            PTAXRow(
                data=d,
                timestamp_ptax_oficial=f"{d}T{fake_ts}",
                cotacao_compra=c,
                cotacao_venda=v,
                usd_brl_mid=round(mid, 4),
            )
        )
    return (
        rows,
        "Fallback público — médias diárias USD-BRL (PoundSterlingLive, secção November 2025)",
    )


def minihash(seed: str) -> str:
    return "mh0x" + hashlib.sha256(seed.encode()).hexdigest()[:56]


def macro_hash(bucket_key: str) -> str:
    return "MH0x" + hashlib.sha256(("MACRO|" + bucket_key).encode()).hexdigest()[:56]


def build_lps(n: int = 100) -> list[LPRow]:
    rng = random.Random(20251105)
    rows: list[LPRow] = []
    for i in range(1, n + 1):
        gen = "F" if rng.random() < 0.52 else "M"
        p1 = rng.choice(PRIMEIROS_F if gen == "F" else PRIMEIROS_M)
        s1 = rng.choice(SOBRENOMES)
        s2 = rng.choice(SOBRENOMES)
        if s2 == s1:
            s2 = SOBRENOMES[(SOBRENOMES.index(s1) + 3) % len(SOBRENOMES)]
        nome = f"{p1} {s1} {s2}"
        city, uf = rng.choice(CIDADES)
        # Regra de Múltiplos de 100 para o Capital Inicial
        cap_inicial = float(rng.randint(50, 500) * 100)
        rows.append(LPRow(
            lp_id=f"LP_{i:03d}", 
            nome_completo=nome, 
            genero=gen, 
            cidade=city, 
            estado=uf,
            capital_inicial=cap_inicial,
            current_capital=cap_inicial
        ))
    return rows


def main() -> None:
    ptax_rows, src = fetch_bcb_ptax()
    if not ptax_rows:
        ptax_rows, src = fallback_ptax()

    by_date: dict[str, PTAXRow] = {}
    for r in ptax_rows:
        by_date[r.data] = r

    OUT_JSON_FX.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON_FX.write_text(
        json.dumps(
            {"fonte": src, "registros": [asdict(x) for x in ptax_rows]},
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    lps = build_lps(100)
    OUT_JSON_LP.write_text(
        json.dumps([asdict(x) for x in lps], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    from collections import deque
    lp_queue = deque(lps)
    
    rng = random.Random(202511)
    tx_rows: list[dict] = []
    lp_share_pct = 0.75

    for day in sorted(by_date.keys()):
        pr = by_date[day]
        rate = pr.usd_brl_mid
        macro_key = f"{day}|PTAX|{pr.timestamp_ptax_oficial}"
        mh_macro = macro_hash(macro_key)
        
        # Simulamos uma demanda de 15 remessas por janela (Manhã/Tarde/Noite)
        for slot, hour in enumerate((9, 13, 17)):
            for ticket in range(15):
                ts_iso = f"{day}T{hour:02d}:{22+ticket:02d}:41-03:00"
                gtv = round(rng.lognormvariate(8, 0.5), 2)
                usd_equiv = round(gtv / rate, 2)
                
                # MOTOR DE FILA (FIFO)
                matched_lp = None
                checked = 0
                while checked < len(lp_queue):
                    potential = lp_queue[0]
                    if potential.current_capital >= gtv:
                        matched_lp = lp_queue.popleft()
                        break
                    else:
                        lp_queue.rotate(-1)
                        checked += 1
                
                if not matched_lp: continue

                spread_pct = get_fluxus_spread_pct(usd_equiv)
                spread_total = round(gtv * spread_pct, 2)
                lp_gain = round(spread_total * lp_share_pct, 2)
                plat_share = round(spread_total - lp_gain, 2)
                
                seed_mini = f"{ts_iso}|{matched_lp.lp_id}|{gtv:.2f}|SETTLED"
                mini = minihash(seed_mini)
                vinculo = f"{mini}→{mh_macro}"
                
                tx_rows.append(
                    {
                        "data_ciclo": day,
                        "timestamp_ciclo_iso": ts_iso,
                        "macro_hash": mh_macro,
                        "minihash": mini,
                        "vinculo_minihash_macrohash": vinculo,
                        "lp_id": matched_lp.lp_id,
                        "lp_nome": matched_lp.nome_completo,
                        "gtv_brl": gtv,
                        "gtv_usd_equiv_ptax_mid": usd_equiv,
                        "spread_total_brl": spread_total,
                        "lp_spread_brl_75pct": lp_gain,
                        "fluxus_orchestration_fee_brl": plat_share,
                        "posicao_na_fila": checked + 1
                    }
                )
                # Atualiza saldo e volta para o fim da fila
                matched_lp.current_capital += lp_gain
                lp_queue.append(matched_lp)

    with OUT_CSV_TX.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(tx_rows[0].keys()))
        w.writeheader()
        w.writerows(tx_rows)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fill_h = PatternFill("solid", fgColor="1B365D")
    font_h = Font(color="FFFFFF", bold=True)
    money_fill = PatternFill("solid", fgColor="FFF2CC")
    hash_fill = PatternFill("solid", fgColor="C6E0B4")

    ws0 = wb.active
    ws0.title = "REFERENCIA"
    rows_ref = [
        ("Cenário temporal", "2026 Harmonizado (Simulação de Fila FIFO; contabilidade PTAX)"),
        ("Fonte USD/BRL", src),
        ("API consultada", "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoDolarPeriodo"),
        ("Spread produto", "Tiered Dinâmico (1,08% - 0,50%) — 75% Share LP"),
        ("DRE agregado", "02_INVESTOR/INVESTOR_DRE_CASE_STUDY_100.md (Share LP 75%)"),
        ("Fila de Fluxo", "Seleção FIFO por saldo integral e alternância de prioridade."),
        ("Minihash", "Prova por ciclo / transição — TECH_STATE_MACHINE_REVERSAL.md"),
        ("Macro Hash", "Agregação temporal de minihashes — TECH_FLUXUS_BLUEPRINT.md"),
    ]
    for i, (a, b) in enumerate(rows_ref, 1):
        ws0.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=b)

    ws1 = wb.create_sheet("PTAX_USD_BRL")
    h1 = ["data", "timestamp_ptax_oficial", "cotacao_compra", "cotacao_venda", "usd_brl_mid"]
    ws1.append(h1)
    for c in ws1[1]:
        c.fill = fill_h
        c.font = font_h
    for r in ptax_rows:
        ws1.append([r.data, r.timestamp_ptax_oficial, r.cotacao_compra, r.cotacao_venda, r.usd_brl_mid])

    ws2 = wb.create_sheet("LP_MASTER")
    ws2.append(["lp_id", "nome_completo", "genero", "cidade", "estado", "capital_inicial"])
    for c in ws2[1]:
        c.fill = fill_h
        c.font = font_h
    for lp in lps:
        ws2.append([lp.lp_id, lp.nome_completo, lp.genero, lp.cidade, lp.estado, lp.capital_inicial])

    ws3 = wb.create_sheet("CICLOS_MINIHASH")
    keys = list(tx_rows[0].keys())
    ws3.append(keys)
    for c in ws3[1]:
        c.fill = fill_h
        c.font = font_h
    highlight_money = {
        "gtv_brl",
        "spread_total_brl_3_5pct",
        "lp_spread_brl_42_8pct",
        "vasp_spread_brl_28_6pct",
        "fluxus_orchestration_brl_28_6pct",
        "usd_brl_ptax_mid_usado",
        "gtv_usd_equiv_ptax_mid",
        "usd_brl_ptax_compra",
        "usd_brl_ptax_venda",
    }
    highlight_hash = {"minihash", "macro_hash", "vinculo_minihash_macrohash", "macro_hash_chave_agregacao"}
    col_idx = {k: i + 1 for i, k in enumerate(keys)}
    for row in tx_rows:
        ws3.append([row[k] for k in keys])
    for name, fl in ((highlight_money, money_fill), (highlight_hash, hash_fill)):
        for k in name:
            if k not in col_idx:
                continue
            L = get_column_letter(col_idx[k])
            for rr in range(2, ws3.max_row + 1):
                ws3[f"{L}{rr}"].fill = fl

    wb.save(OUT_XLSX)
    print("OK", OUT_XLSX)


if __name__ == "__main__":
    main()
